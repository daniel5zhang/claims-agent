"""
数据库初始化脚本
解析 Excel 素材库、历史方案、docx 服务手册模板，导入 seekdb/SQLite
"""
import os
import sys
import json
from decimal import Decimal
from pathlib import Path

# 将 backend 加入路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from database import init_db, SessionLocal
from database import Service, Scheme, SchemeItem, ServiceManualTemplate
from services.manual_generator import parse_manual_template


DOC_DIR = Path("/Users/daniel/Desktop/code/产品精算/doc")
DATA_DIR = Path(__file__).parent


def init_all():
    """初始化所有数据"""
    print("=== 初始化数据库 ===")
    init_db()
    db = SessionLocal()
    try:
        print(">>> 导入服务素材...")
        import_services(db)
        print(">>> 导入历史方案...")
        import_historical_schemes(db)
        print(">>> 导入服务手册模板...")
        import_manual_templates(db)
        print(">>> 完成！")
    finally:
        db.close()


def _parse_price(val):
    """安全解析价格，处理数字、范围、文本混合的情况"""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    if isinstance(val, str):
        s = val.strip().replace(',', '')
        # 尝试提取第一个数字
        import re
        m = re.search(r'\d+(?:\.\d+)?', s)
        if m:
            return Decimal(m.group())
    return None


def _parse_rate(val):
    """安全解析使用率/百分比"""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    if isinstance(val, str):
        s = val.strip().replace('%', '')
        try:
            return Decimal(s)
        except:
            pass
    return None


def import_services(db):
    """从 Excel 导入服务素材"""
    try:
        import openpyxl
    except ImportError:
        print("警告: 未安装 openpyxl，跳过服务素材导入")
        return

    file_path = DOC_DIR / "产品精算中心健管服务明细一览表.xlsx"
    if not file_path.exists():
        print(f"警告: 素材文件不存在: {file_path}")
        # 创建一些示例数据以便开发测试
        _create_sample_services(db)
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # === 产品部 sheet ===
    # 列结构: 服务类别(0), 服务项目(1), 服务说明(2), 服务流程(3), 启动条件(4), 服务次数(5), 实发成本价(6), 跟单发生率\使用率5万单以下(7), 跟单发生率\使用率5-10万(8)
    if "产品部" in wb.sheetnames:
        ws = wb["产品部"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            svc = Service(
                category=str(row[0]) if row[0] else None,
                name=str(row[1]) if row[1] else "未命名",
                description=str(row[2]) if len(row) > 2 else None,
                process=str(row[3]) if len(row) > 3 else None,
                condition=str(row[4]) if len(row) > 4 else None,
                times=str(row[5]) if len(row) > 5 else None,
                cost_price=_parse_price(row[6]) if len(row) > 6 else None,
                usage_rate_small=_parse_rate(row[7]) if len(row) > 7 else None,
                usage_rate_large=_parse_rate(row[8]) if len(row) > 8 else None,
                source_file="产品精算中心健管服务明细一览表.xlsx",
            )
            db.add(svc)

    # === 交付部 sheet ===
    # 列结构: 服务类别(0), 服务项目(1), 服务说明(2), 服务流程(3), 启动条件(4), 服务次数(5)
    if "交付部" in wb.sheetnames:
        ws = wb["交付部"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            svc = Service(
                category=str(row[0]) if row[0] else "交付部",
                name=str(row[1]) if row[1] else "未命名",
                description=str(row[2]) if len(row) > 2 else None,
                process=str(row[3]) if len(row) > 3 else None,
                condition=str(row[4]) if len(row) > 4 else None,
                times=str(row[5]) if len(row) > 5 else None,
                source_file="产品精算中心健管服务明细一览表.xlsx",
            )
            db.add(svc)

    db.commit()
    count = db.query(Service).count()
    print(f"  已导入 {count} 条服务素材")


def import_historical_schemes(db):
    """从历史方案 Excel 导入"""
    try:
        import openpyxl
    except ImportError:
        print("警告: 未安装 openpyxl，跳过历史方案导入")
        return

    scheme_files = [
        "中荷重疾绿通服务升级版方案（20260324内部报价）.xlsx",
        "海峡随车健康管理服务四个方案报价-4.7（内部分项报价）.xlsx",
        "海峡随车健康管理服务四个方案报价-4.10（外部报价反馈）.xlsx",
    ]

    total = 0
    for filename in scheme_files:
        file_path = DOC_DIR / filename
        if not file_path.exists():
            print(f"  跳过（文件不存在）: {filename}")
            continue
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scheme_name = f"{filename.split('.')[0]}_{sheet_name}"
                scheme = Scheme(
                    scheme_name=scheme_name[:200],
                    customer_name=None,
                    version="1.0",
                    source_file=filename,
                )
                db.add(scheme)
                db.flush()

                services = []
                for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                    if not row:
                        continue
                    row_text = " ".join(str(c) for c in row if c)
                    if "服务" in row_text and len(row) >= 2:
                        services.append({
                            "name": str(row[0]) if row[0] else "",
                            "detail": row_text,
                        })

                if services:
                    scheme.service_list_json = json.dumps(services, ensure_ascii=False)
                    total += 1

            db.commit()
        except Exception as e:
            print(f"  解析失败 {filename}: {e}")
            db.rollback()

    # 国任财险文件是 .xls 格式，用 pandas 读取
    guoren_file = DOC_DIR / "国任财险-河北城商行渠道健康管理服务20260316.xls"
    if guoren_file.exists():
        try:
            import pandas as pd
            df = pd.read_excel(guoren_file, sheet_name=None)
            for sheet_name in df.keys():
                scheme_name = f"国任财险-河北城商行渠道健康管理服务20260316_{sheet_name}"
                scheme = Scheme(
                    scheme_name=scheme_name[:200],
                    source_file="国任财险-河北城商行渠道健康管理服务20260316.xls",
                )
                db.add(scheme)
                db.flush()
            db.commit()
            total += len(df)
        except Exception as e:
            print(f"  解析失败 国任财险: {e}")
            db.rollback()

    print(f"  已导入 {total} 个历史方案")


def import_manual_templates(db):
    """从 docx 服务手册导入模板"""
    file_path = DOC_DIR / "8.职工家庭防癌抗癌保障卡健康管理服务手册（个人尊享版）.docx"
    if not file_path.exists():
        print(f"警告: 服务手册文件不存在: {file_path}")
        return

    try:
        templates = parse_manual_template(str(file_path))
        for t in templates:
            template = ServiceManualTemplate(
                template_name=t["template_name"],
                service_name=t["service_name"],
                section_title=t["section_title"],
                content_template=t["content_template"],
                sort_order=t["sort_order"],
            )
            db.add(template)
        db.commit()
        count = db.query(ServiceManualTemplate).count()
        print(f"  已导入 {count} 条服务手册模板")
    except Exception as e:
        print(f"  解析失败: {e}")
        db.rollback()


def _create_sample_services(db):
    """创建示例服务数据（开发调试用）"""
    sample_services = [
        {"category": "基础服务", "name": "图文问诊", "description": "7x24小时在线图文咨询", "times": "不限次", "cost_price": "5.00"},
        {"category": "基础服务", "name": "视频问诊", "description": "三甲医院专家视频咨询", "times": "2次/年", "cost_price": "30.00"},
        {"category": "基础服务", "name": "电话心理咨询", "description": "专业心理咨询师电话服务", "times": "1次/年", "cost_price": "20.00"},
        {"category": "重疾服务", "name": "重疾门诊绿通", "description": "协助预约三甲医院专家门诊", "times": "1次", "cost_price": "80.00"},
        {"category": "重疾服务", "name": "就医陪诊服务", "description": "专业陪诊人员陪同就医", "times": "1次", "cost_price": "50.00"},
        {"category": "重疾服务", "name": "重疾住院护理服务", "description": "5天4晚专业住院陪护", "times": "2次", "cost_price": "150.00"},
        {"category": "重疾服务", "name": "肿瘤专家会诊优惠服务", "description": "线上专家会诊优惠", "times": "1次", "cost_price": "200.00"},
        {"category": "重疾服务", "name": "质子重离子就医直通车", "description": "质子重离子医院就医服务", "times": "1次", "cost_price": "100.00"},
        {"category": "药品服务", "name": "购药金", "description": "在线购药抵扣金", "times": "按需", "cost_price": "10.00"},
        {"category": "药品服务", "name": "在线药品商城", "description": "优惠购药平台", "times": "不限次", "cost_price": "5.00"},
    ]
    for s in sample_services:
        svc = Service(
            category=s["category"],
            name=s["name"],
            description=s["description"],
            times=s["times"],
            cost_price=Decimal(s["cost_price"]) if s.get("cost_price") else None,
            source_file="sample",
        )
        db.add(svc)
    db.commit()
    print(f"  已创建 {len(sample_services)} 条示例服务素材")


if __name__ == "__main__":
    init_all()
