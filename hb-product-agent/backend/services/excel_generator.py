"""
Excel 报价单生成器
生成两个 Sheet：
  Sheet1 - 服务清单（参考海峡随车Excel的"健康管理服务"sheet）
  Sheet2 - 方案内容（参考海峡随车Excel的"方案"sheet）
"""
import json
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import GeneratedScheme, GeneratedExcel, Service


class ExcelGenerator:
    """生成方案报价单 Excel"""

    def generate_excel(self, db: Session, scheme_id: int):
        scheme = db.query(GeneratedScheme).filter(GeneratedScheme.id == scheme_id).first()
        if not scheme:
            raise ValueError("方案不存在")

        # 解析服务清单（健壮解析：兼容 dict/list/str 多种格式）
        service_list = []
        schemes_list = []  # 多方案数据
        if scheme.service_list_json:
            try:
                parsed = json.loads(scheme.service_list_json)
                if isinstance(parsed, dict):
                    service_list = parsed.get("services", []) or []
                    schemes_list = parsed.get("schemes", []) or []
                elif isinstance(parsed, list):
                    service_list = parsed
            except (json.JSONDecodeError, TypeError):
                service_list = []

        # 确保 service_list 是 list 且每个元素是 dict
        service_list = [s for s in service_list if isinstance(s, dict)]
        # 确保 schemes_list 中每个方案的 services 是有效 list
        for sch in schemes_list:
            if isinstance(sch, dict):
                if "services" in sch:
                    sch["services"] = [s for s in sch["services"] if isinstance(s, dict)]
                elif "service_list" in sch:
                    # 兼容 LLM 可能使用 service_list 而非 services 的情况
                    sch["services"] = [s for s in sch["service_list"] if isinstance(s, dict)]

        # 修复：如果第一个方案的 services 为空但顶层 service_list 有数据，则补全
        if schemes_list and isinstance(schemes_list[0], dict):
            first_svc = schemes_list[0].get("services", [])
            if not first_svc and service_list:
                schemes_list[0]["services"] = service_list

        # 汇总所有方案的服务用于服务清单 Sheet（按方案分组展示）
        all_services_with_scheme = []
        if schemes_list:
            for sch in schemes_list:
                if isinstance(sch, dict):
                    sch_name = sch.get("scheme_name", "")
                    for svc in sch.get("services", []):
                        if isinstance(svc, dict):
                            all_services_with_scheme.append({"scheme": sch_name, **svc})
        if not all_services_with_scheme:
            # 如果 schemes 为空，用顶层 service_list（单方案模式）
            all_services_with_scheme = [{"scheme": "", **s} for s in service_list]

        # 查询素材库补充详细信息
        db_services = {s.name: s for s in db.query(Service).all()}

        # 创建 Excel
        wb = Workbook()

        # ========== Sheet1: 服务清单 ==========
        ws1 = wb.active
        ws1.title = "服务清单"

        # 样式
        title_font = Font(name="微软雅黑", size=16, bold=True, color="003366")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        normal_font = Font(name="微软雅黑", size=10)
        bold_font = Font(name="微软雅黑", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 标题
        ws1.merge_cells("A1:I1")
        ws1["A1"] = f"{scheme.scheme_name or '产品方案'} - 服务清单"
        ws1["A1"].font = title_font
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 30

        # 表头（参考海峡随车Excel Sheet1）
        # 多方案时增加"方案档位"列，单方案时不加
        has_multi = any(s.get("scheme") for s in all_services_with_scheme)
        headers1 = ["序号"]
        if has_multi:
            headers1.append("方案档位")
        headers1.extend(["服务项目", "服务内容", "服务次数", "启动条件", "服务标准", "服务网络", "成本价（元）"])
        table_start = 3
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=table_start, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws1.row_dimensions[table_start].height = 24

        # 数据
        col_offset = 1 if has_multi else 0  # 方案档位列偏移
        for idx, svc in enumerate(all_services_with_scheme, 1):
            if not isinstance(svc, dict):
                continue
            row = table_start + idx
            svc_name = svc.get("name", "")
            db_svc = None
            for db_name, db_s in db_services.items():
                if svc_name in db_name or db_name in svc_name:
                    db_svc = db_s
                    break

            col = 1
            ws1.cell(row=row, column=col, value=idx).alignment = center_align
            col += 1
            if has_multi:
                ws1.cell(row=row, column=col, value=svc.get("scheme", "")).alignment = center_align
                col += 1
            ws1.cell(row=row, column=col, value=svc_name).alignment = left_align
            col += 1
            ws1.cell(row=row, column=col, value=db_svc.description if db_svc and db_svc.description else "").alignment = left_align
            col += 1
            ws1.cell(row=row, column=col, value=svc.get("times", db_svc.times if db_svc and db_svc.times else "")).alignment = center_align
            col += 1
            ws1.cell(row=row, column=col, value=svc.get("condition", db_svc.condition if db_svc and db_svc.condition else "")).alignment = left_align
            col += 1
            ws1.cell(row=row, column=col, value=db_svc.service_standard if db_svc and db_svc.service_standard else "").alignment = left_align
            col += 1
            ws1.cell(row=row, column=col, value=db_svc.service_network if db_svc and db_svc.service_network else "").alignment = left_align
            col += 1
            cost = svc.get("cost_price") or svc.get("cost", "")
            ws1.cell(row=row, column=col, value=cost).alignment = center_align

            max_col = 8 + col_offset
            for c in range(1, max_col + 1):
                ws1.cell(row=row, column=c).border = thin_border
                ws1.cell(row=row, column=c).font = normal_font

        # 列宽
        ws1.column_dimensions["A"].width = 8
        if has_multi:
            ws1.column_dimensions["B"].width = 14
        name_col = get_column_letter(2 + col_offset)
        ws1.column_dimensions[name_col].width = 22
        content_col = get_column_letter(3 + col_offset)
        ws1.column_dimensions[content_col].width = 40
        times_col = get_column_letter(4 + col_offset)
        ws1.column_dimensions[times_col].width = 12
        cond_col = get_column_letter(5 + col_offset)
        ws1.column_dimensions[cond_col].width = 20
        std_col = get_column_letter(6 + col_offset)
        ws1.column_dimensions[std_col].width = 25
        net_col = get_column_letter(7 + col_offset)
        ws1.column_dimensions[net_col].width = 25
        cost_col = get_column_letter(8 + col_offset)
        ws1.column_dimensions[cost_col].width = 14

        # ========== 方案内容 Sheet（支持多方案） ==========
        if schemes_list:
            # 多方案：每个方案一个 Sheet
            for sch_idx, sch in enumerate(schemes_list):
                sch_name = sch.get("scheme_name", f"方案{sch_idx + 1}")
                # Sheet 名称最长31字符，且不能含 \ / * ? [ ] :
                sheet_title = re.sub(r"[\\/*?:\[\]]", "", sch_name)
                sheet_title = sheet_title[:28] + "..." if len(sheet_title) > 31 else sheet_title
                self._write_scheme_sheet(
                    wb, sheet_title, sch_name,
                    scene=sch.get("scene", scheme.scene or "-"),
                    target_group=sch.get("target_group", scheme.target_group or "-"),
                    total_cost=sch.get("total_cost", scheme.total_cost),
                    total_quote=sch.get("total_quote", scheme.total_quote),
                    service_list=sch.get("services", []),
                    title_font=title_font, bold_font=bold_font,
                    header_font=header_font, header_fill=header_fill,
                    normal_font=normal_font, center_align=center_align,
                    left_align=left_align, thin_border=thin_border,
                )
        else:
            # 单方案
            self._write_scheme_sheet(
                wb, "方案内容", scheme.scheme_name or '产品方案',
                scene=scheme.scene or "-",
                target_group=scheme.target_group or "-",
                total_cost=scheme.total_cost,
                total_quote=scheme.total_quote,
                service_list=service_list,
                title_font=title_font, bold_font=bold_font,
                header_font=header_font, header_fill=header_fill,
                normal_font=normal_font, center_align=center_align,
                left_align=left_align, thin_border=thin_border,
            )

        # 计算版本号
        existing_count = db.query(GeneratedExcel).filter(
            GeneratedExcel.scheme_id == scheme_id
        ).count()
        version = existing_count + 1

        # 安全文件名
        safe_name = scheme.scheme_name or "方案"
        safe_name = safe_name.replace("/", "_").replace("\\", "_")[:60]

        output_dir = os.path.join(os.path.dirname(__file__), "..", "..", "output", "excels")
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)
        filename = f"{safe_name}_v{version}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)

        # 保存到数据库
        excel = GeneratedExcel(
            scheme_id=scheme_id,
            version=version,
            excel_title=scheme.scheme_name,
            excel_path=filepath,
            status="generated",
        )
        db.add(excel)
        db.flush()
        db.refresh(excel)

        return excel

    def _write_scheme_sheet(
        self, wb, sheet_title, scheme_name, scene, target_group,
        total_cost, total_quote, service_list,
        title_font, bold_font, header_font, header_fill,
        normal_font, center_align, left_align, thin_border,
    ):
        """写入单个方案内容 Sheet"""
        ws = wb.create_sheet(title=sheet_title)

        # 标题
        ws.merge_cells("A1:E1")
        ws["A1"] = f"{scheme_name} - 方案内容"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 方案说明
        ws["A3"] = "方案名称"
        ws["B3"] = scheme_name or "-"
        ws["A4"] = "适用场景"
        ws["B4"] = scene or "-"
        ws["A5"] = "目标人群"
        ws["B5"] = target_group or "-"
        ws["A6"] = "总成本"
        ws["B6"] = f"{total_cost} 元/人/年" if total_cost else "-"
        ws["A7"] = "总报价"
        ws["B7"] = f"{total_quote} 元/人/年" if total_quote else "-"
        for r in [3, 4, 5, 6, 7]:
            ws[f"A{r}"].font = bold_font
            ws[f"A{r}"].alignment = left_align

        # 表头
        headers = ["序号", "服务项目", "服务次数", "成本价（元）", "报价（元）"]
        table_start = 9
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[table_start].height = 24

        # 数据
        for idx, svc in enumerate(service_list, 1):
            if not isinstance(svc, dict):
                continue
            row = table_start + idx
            ws.cell(row=row, column=1, value=idx).alignment = center_align
            ws.cell(row=row, column=2, value=svc.get("name", "")).alignment = left_align
            ws.cell(row=row, column=3, value=svc.get("times", "")).alignment = center_align
            cost = svc.get("cost_price") or svc.get("cost", "")
            quote = svc.get("quote_price") or svc.get("price") or svc.get("quote", "")
            ws.cell(row=row, column=4, value=cost).alignment = center_align
            ws.cell(row=row, column=5, value=quote).alignment = center_align
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).font = normal_font

        # 合计行
        total_row = table_start + len(service_list) + 1
        ws.cell(row=total_row, column=1, value="合计").alignment = center_align
        ws.cell(row=total_row, column=2, value="").alignment = center_align
        ws.cell(row=total_row, column=3, value="").alignment = center_align
        try:
            cost_val = float(total_cost) if total_cost else 0
        except (ValueError, TypeError):
            cost_val = 0
        try:
            quote_val = float(total_quote) if total_quote else 0
        except (ValueError, TypeError):
            quote_val = 0
        ws.cell(row=total_row, column=4, value=cost_val).alignment = center_align
        ws.cell(row=total_row, column=5, value=quote_val).alignment = center_align
        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border
            cell.font = bold_font
            cell.fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")

        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

        # 底部说明
        footer_row = total_row + 2
        ws.merge_cells(f"A{footer_row}:E{footer_row}")
        ws[f"A{footer_row}"] = "说明：以上报价为圆心惠保内部成本及对外报价参考，具体商务条款以最终合同为准。"
        ws[f"A{footer_row}"].font = Font(name="微软雅黑", size=9, italic=True, color="909399")
        ws[f"A{footer_row}"].alignment = left_align
